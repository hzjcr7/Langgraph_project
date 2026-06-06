from __future__ import annotations

from pathlib import Path
import sys
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.agent_finance_automation.config import get_settings
from src.agent_finance_automation.evaluation import run_evaluation


def main() -> int:
    settings = get_settings()
    report = run_evaluation(settings)

    excel_path = Path(report["artifacts"]["bank"]["excel"])
    docx_path = Path(report["artifacts"]["regulatory"]["docx"])

    workbook = load_workbook(excel_path)
    assert workbook.active["A1"].value == "Account Name"
    assert workbook.active.max_row >= 12

    with ZipFile(docx_path) as docx:
        assert "word/document.xml" in docx.namelist()
        document_xml = docx.read("word/document.xml").decode("utf-8")
        assert "SFC Regulatory Digest" in document_xml

    assert report["compliance"]["high_risk_regulatory_items"] >= 1
    print("smoke test passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
